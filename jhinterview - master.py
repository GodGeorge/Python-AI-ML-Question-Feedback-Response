from openpyxl import load_workbook
import openpyxl

def SearchCell(search_str, range=None):

    global ws

    range = ws.iter_rows()  # Defaults to whole sheet

    for tupleOfCells in range:
        for cell in tupleOfCells:
            if (cell.value == search_str):
                return [_tuple[0] for _tuple in ws.iter_cols(min_row=cell.row, max_row=cell.row)]


wb = openpyxl.load_workbook("jhintentwrap.xlsx")
ws = wb.worksheets[0]

search_str = input("Please write the question? ")

# Search only Column 'C' == 3, openpyxl is 1-based
cellsOfFoundRow = SearchCell(search_str, ws.iter_rows(min_col=2, max_col=2))

if cellsOfFoundRow:
    print("Question: ",search_str)
    print("Intent: " ,(cellsOfFoundRow[0].value))
    print("Priority: Low")
    print("Response : A balanced diet is key for any growth of a baby")

else:
    print("Could not find '{}' in the given cell range!".format(search_str))
    print("We have noted your question and we will be providing a solution soon")